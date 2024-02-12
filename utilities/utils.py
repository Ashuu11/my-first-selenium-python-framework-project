import csv

import softest
import logging
import inspect
from openpyxl import load_workbook


class Utils(softest.TestCase):
    def assertListItem(self, list_elem, value):
        for stop in list_elem:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("Assertion passed")
            else:
                print("Assertion Failed")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # Set class name where from where its called
        logger_name = inspect.stack()[1][3]
        # Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create the console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log")
        # create formatter - how we want our logs to be
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s : %(message)s",
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler in logger
        logger.addHandler(fh)
        return logger

    def read_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_from_csv(file_name):
        # Create an empty list
        datalist = []
        # Open the csv file
        csv_data = open(file_name, "r")
        # Create CSV Reader
        reader = csv.reader(csv_data)
        # Skip the header
        next(reader)
        # Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist


